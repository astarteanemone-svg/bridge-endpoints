import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl.styles import PatternFill

OVERPASS = "https://lz4.overpass-api.de/api/interpreter"

def decimal_to_dms(lat, lon):
    def conv(v, lat=True):
        d = 'N' if lat and v >= 0 else 'S' if lat else ('E' if v >= 0 else 'W')
        v = abs(v)
        deg = int(v); m = int((v - deg) * 60); s = (v - deg - m/60) * 3600
        return f"{deg}°{m}'{s:.1f}\"{d}"
    return conv(lat, True), conv(lon, False)

def get_way_and_endpoints(name, area_id):
    query = f"""
    [out:json][timeout:60];
    area({area_id})->.a;
    way["bridge"="yes"]["name"~"^{name}$"](area.a);
    out body;
    >;
    out skel qt;
    """
    r = requests.get(OVERPASS, params={"data": query})
    if r.status_code != 200:
        return None, None, None
    js = r.json()
    nodes = {e["id"]:(e.get("lat"), e.get("lon")) for e in js.get("elements",[]) if e["type"]=="node"}
    ways = [e for e in js.get("elements",[]) if e["type"]=="way"]
    if not ways:
        return None, None, None
    w = ways[0]
    nids = w.get("nodes", [])
    if not nids: return None, None, None
    return w["id"], nodes.get(nids[0]), nodes.get(nids[-1])

def get_nameless_bridges(area_id):
    query = f"""
    [out:json][timeout:60];
    area({area_id})->.a;
    way["bridge"="yes"]["name"!~"."](area.a);
    out body;
    >;
    out skel qt;
    """
    r = requests.get(OVERPASS, params={"data": query})
    if r.status_code != 200:
        return []
    js = r.json()
    nodes = {e["id"]:(e.get("lat"), e.get("lon")) for e in js.get("elements",[]) if e["type"]=="node"}
    ways = [e for e in js.get("elements",[]) if e["type"]=="way"]
    results = []
    for w in ways:
        nids = w.get("nodes", [])
        if not nids: continue
        s = nodes.get(nids[0]); e = nodes.get(nids[-1])
        if not s or not e: continue
        slat, slon = s; elat, elon = e
        sdms_lat, sdms_lon = decimal_to_dms(slat, slon)
        edms_lat, edms_lon = decimal_to_dms(elat, elon)
        results.append({
            "橋名": "橋名なし候補",
            "AreaID": area_id,
            "way_id": f"https://www.openstreetmap.org/way/{w['id']}",
            "起点_緯度(十進)": slat, "起点_経度(十進)": slon,
            "終点_緯度(十進)": elat, "終点_経度(十進)": elon,
            "起点_緯度(度分秒)": sdms_lat, "起点_経度(度分秒)": sdms_lon,
            "終点_緯度(度分秒)": edms_lat, "終点_経度(度分秒)": edms_lon
        })
    return results

st.set_page_config(page_title="橋リスト → 緯度経度変換", page_icon="🌉", layout="wide")
st.title("橋リスト → 緯度経度変換ツール（OSM / Overpass）")

uploaded = st.file_uploader("橋リスト.xlsx を選択してください", type=["xlsx"])

if uploaded:
    df = pd.read_excel(uploaded, sheet_name="橋リスト")

    st.subheader("📂 アップロードした橋リスト")
    st.dataframe(df, use_container_width=True)

    success_rows, failed_rows, candidate_rows = [], [], []
    progress = st.progress(0)

    for i, row in df.iterrows():
        name = str(row["橋名"]).strip() if pd.notna(row["橋名"]) else ""
        area_id = row["AreaID"]

        # 「橋名なし」と明示された場合 → 候補検索
        if name == "橋名なし" and pd.notna(area_id):
            candidates = get_nameless_bridges(int(area_id))
            for c in candidates:
                candidate_rows.append({
                    "橋名": c["橋名"],
                    "県名": row["県名"],
                    "市町村": row["市町村"],
                    "AreaID": c["AreaID"],
                    "way_id": c["way_id"],
                    "起点_緯度(十進)": c["起点_緯度(十進)"],
                    "起点_経度(十進)": c["起点_経度(十進)"],
                    "終点_緯度(十進)": c["終点_緯度(十進)"],
                    "終点_経度(十進)": c["終点_経度(十進)"],
                    "起点_緯度(度分秒)": c["起点_緯度(度分秒)"],
                    "起点_経度(度分秒)": c["起点_経度(度分秒)"],
                    "終点_緯度(度分秒)": c["終点_緯度(度分秒)"],
                    "終点_経度(度分秒)": c["終点_経度(度分秒)"],
                })
            progress.progress((i+1)/len(df))
            continue

        # 橋名が空の場合は未ヒット
        if not name:
            failed_rows.append({"橋名": name, "県名": row["県名"], "市町村": row["市町村"], "AreaID": area_id})
            progress.progress((i+1)/len(df))
            continue

        if pd.isna(area_id):
            failed_rows.append({"橋名": name, "県名": row["県名"], "市町村": row["市町村"], "AreaID": area_id})
            progress.progress((i+1)/len(df))
            continue

        # 通常の橋検索
        way_id, s, e = get_way_and_endpoints(name, int(area_id))
        if way_id and s and e and all(s) and all(e):
            slat, slon = s
            elat, elon = e
            sdms_lat, sdms_lon = decimal_to_dms(slat, slon)
            edms_lat, edms_lon = decimal_to_dms(elat, elon)

            success_rows.append({
                "橋名": name,
                "県名": row["県名"],
                "市町村": row["市町村"],
                "AreaID": area_id,
                "way_id": f"https://www.openstreetmap.org/way/{way_id}",
                "起点_緯度(十進)": slat, "起点_経度(十進)": slon,
                "終点_緯度(十進)": elat, "終点_経度(十進)": elon,
                "起点_緯度(度分秒)": sdms_lat, "起点_経度(度分秒)": sdms_lon,
                "終点_緯度(度分秒)": edms_lat, "終点_経度(度分秒)": edms_lon
            })
        else:
            failed_rows.append({"橋名": name, "県名": row["県名"], "市町村": row["市町村"], "AreaID": area_id})

        progress.progress((i+1)/len(df))

    df_success = pd.DataFrame(success_rows)
    df_failed  = pd.DataFrame(failed_rows)
    df_candidates = pd.DataFrame(candidate_rows)

    if not df_success.empty:
        st.subheader("✅ 成功した橋（プレビュー）")
        st.dataframe(df_success.head(50), use_container_width=True)

    if not df_failed.empty:
        st.subheader("⚠️ 未ヒット橋（OSMで見つからず）")
        st.dataframe(df_failed, use_container_width=True)

    if not df_candidates.empty:
        st.subheader("🔎 橋名なし候補")
        st.dataframe(df_candidates.head(50), use_container_width=True)

    # Excel出力
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not df_success.empty:
            df_success.to_excel(writer, sheet_name="成功した橋", index=False)
        if not df_failed.empty:
            df_failed.to_excel(writer, sheet_name="未ヒット橋", index=False)
        if not df_candidates.empty:
            df_candidates.to_excel(writer, sheet_name="橋名なし候補", index=False)

        wb = writer.book

        # 成功した橋シートの加工
        if "成功した橋" in wb.sheetnames:
            ws = wb["成功した橋"]
            blue   = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
            orange = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

            headers = [c.value for c in ws[1]]
            for col_idx, h in enumerate(headers, start=1):
                if not h: continue
                if str(h).startswith("起点"):
                    for r in range(2, ws.max_row+1):
                        ws.cell(row=r, column=col_idx).fill = blue
                elif str(h).startswith("終点"):
                    for r in range(2, ws.max_row+1):
                        ws.cell(row=r, column=col_idx).fill = orange

            if "way_id" in headers:
                col_idx = headers.index("way_id") + 1
                for r in range(2, ws.max_row+1):
                    url = ws.cell(row=r, column=col_idx).value
                    if url:
                        ws.cell(row=r, column=col_idx).hyperlink = url
                        ws.cell(row=r, column=col_idx).style = "Hyperlink"

        # 橋名なし候補シートの加工
        if "橋名なし候補" in wb.sheetnames:
            ws = wb["橋名なし候補"]
            headers = [c.value for c in ws[1]]
            if "way_id" in headers:
                col_idx = headers.index("way_id") + 1
                for r in range(2, ws.max_row+1):
                    url = ws.cell(row=r, column=col_idx).value
                    if url:
                        ws.cell(row=r, column=col_idx).hyperlink = url
                        ws.cell(row=r, column=col_idx).style = "Hyperlink"

    st.download_button(
        label="📥 結果をダウンロード（Excel：bridge_endpoints.xlsx）",
        data=output.getvalue(),
        file_name="bridge_endpoints.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
